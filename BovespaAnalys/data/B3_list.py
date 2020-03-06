'''
Author: Mauro Matsudo
This script uses the official B3 web site get the list of all thhe firm trade in brazilian stock market
'''


if __name__ == "__main__":

    import openpyxl
    from pandas import DataFrame
    from requests import get
    from zipfile import ZipFile
    from io import BytesIO
    from sys import exit
    from os.path import exists

    request = get('http://www.b3.com.br/lumis/portal/file/fileDownload.jsp?fileId=8AA8D0975A2D7918015A3C81693D4CA4')
    if request.status_code is False:
        print('An Error occurred, if evertything alright with your internet and proxy, so the worksheet is no longer available in '
              'http://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/acoes/consultas/classificacao-setorial/')
        exit(request.status_code)

    file = ZipFile(BytesIO(request.content))
    name = file.infolist()[0].filename

    # The file name contain release date, so if the worksheet was previously downloaded and has the same date
    # it isn't necessay to upgrade our data
    if exists(name):
        print("The excel file from B3 is the  last release. There is no reason to download new version.\n"
              "If there is any problem with your plan, please delete it and run the scrip again!")
        exit(0)

    file.extractall()
    b3_plan = openpyxl.load_workbook(name)

    # It'll select the first plan doesn't matter its name, so it b3 change it, there'll be no effect
    sheet = b3_plan[b3_plan.sheetnames[0]]
    # Nornally the column D is responsable to store the ticker, so it will be our reference
    max_row_d = max((d.row for d in sheet['D'] if d.value is not None)) # Get the number of companies trade in B3

    tickers = {}
    for row in sheet.iter_rows(min_row=1, max_row= max_row_d, min_col=4 ,max_col=4): # iterating the rows with the tickers
        current_row = row[0]
        ticker = current_row.value
        industry_cell = sheet.cell(row=current_row.row, column=1)  # the instustry sector is stored in the first column
        if  industry_cell.value == 'SETOR ECONÔMICO':
            # the general industry is defined bellow the 'SETOR ECONÔMICO' header, however that header his merged and
            # occupies 2 rows, that's why there +2. Note that, until the next header, all the firms belongs to the
            # same industry
            industry = sheet.cell(row=(industry_cell.row+2), column=1).value
        if ticker != None and (len(ticker) == 4):
            #tickers.append(ticker)
            row_addr = current_row.row
            tickers[row_addr] = {'Ticker': ticker,
                                'Trade Name': sheet.cell(row=row_addr, column=3).value.strip(),
                                 'Industry': industry.strip()}
    b3_df = DataFrame.from_dict(tickers, orient='index')
    b3_df.to_excel("B3_list.xlsx", index=False)


'''from requests import get
from bs4 import BeautifulSoup
from sys import exit

#companies_list = get('http://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/empresas-listadas.htm')

# source: https://www.guiainvest.com.br/lista-acoes/default.aspx?listaacaopage=1



def convert_to_table(html, css_selector):
    html = BeautifulSoup(html, 'html.parser')
    return html.select(css_selector)[0]


html = get('https://www.guiainvest.com.br/lista-acoes/default.aspx?listaacaopage=1').text

text = BeautifulSoup(html, 'html.parser')
pages = text.select('div.rgWrap:nth-child(2)')
#pages = pages.find_all('span')
#pages = [page.text for page in pages]

#print(convert_to_table(html, '#RadGrid1 > table > tbody'))
table = BeautifulSoup(html, 'html.parser')
table = table.find('tbody')
companies = table.find_all('td')[0]
#[element.text for element in html_table.find_all('h3')]

# remove the logo image from the html table
for photo in companies.find_all('a', {'class':'memberPhoto'}):
    photo.decompose()
tickers = [element.text for element in companies.find_all('a')]

#the ticker name was inside of the a tag, so it was previouslys got, before remove all the links
for link in companies.find_all('a'):
    link.decompose()

companies = companies.get_text(separator = '\n').split('\n')
full_name = [name for name in companies if companies.index(name)%2 == 0]
industry = [name for name in companies if companies.index(name)%2 != 0]

if len(tickers) == len(full_name) and len(full_name) == len(industry):
    pass
else:
    print("The code to get the companies list is no longer correct. There are differences in the rows of the table,"
          " some columns has more rows than others. The html site code have been modifyed")
    exit(1)

from pandas import DataFrame
companies_list = DataFrame(columns = ["Ticker", "Trade name", "Industry"])
companies_list["Ticker"] = tickers
companies_list["Trade name"] = full_name
companies_list["Industry"] = industry
print(companies_list)'''