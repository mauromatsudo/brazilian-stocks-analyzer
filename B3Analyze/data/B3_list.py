'''
Author: Mauro Matsudo
This script uses the official B3 web site get the list of all thhe firm trade in brazilian stock market
'''

import openpyxl
import requests
from pandas import DataFrame
from zipfile import ZipFile
from io import BytesIO
from sys import exit
from os.path import exists

class Plan:
    def __init__(self):
        self._url = 'http://www.b3.com.br/lumis/portal/file/fileDownload.jsp?fileId=8AA8D0975A2D7918015A3C81693D4CA4'

    def download_plan(self):
        try:
            request = requests.get(self._url)
        except requests.exceptions.ConnectionError:
            print('There is a problem with your connection to http://www.b3.com.br.')
        if request.status_code is False:
            print(
                'An Error occurred, if evertything alright with your internet and proxy, so the worksheet is no longer available in '
                'http://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/acoes/consultas/classificacao-setorial/')
            exit(request.status_code)
        file = ZipFile(BytesIO(request.content))
        name = file.infolist()[0].filename
        # The file name contain release date, so if the worksheet was previously downloaded and has the same date
        # it isn't necessay to upgrade our data
        if exists(name):
            print("The excel file from B3 is the  last release. There is no reason to download new version.\n"
                  "If there is any problem with your plan, please delete it and run the scrip again!")
        else:
            file.extractall()
        return openpyxl.load_workbook(name)

    def organize_plan(self, download_new=False):
        if download_new == True:
            plan = self.download_plan()
        else:
            plan = openpyxl.load_workbook("Setorial B3 03-03-2020 (português).xlsx")
        # It'll select the first plan doesn't matter its name, so it b3 change it, there'll be no effect
        sheet = plan[plan.sheetnames[0]]
        # Nornally the column D is responsable to store the ticker, so it will be our reference
        max_row_d = max((d.row for d in sheet['D'] if d.value is not None)) # Get the number of companies trade in B3
        tickers = {}
        for row in sheet.iter_rows(min_row=1, max_row= max_row_d, min_col=4 ,max_col=4): # iterating the rows with the tickers
            current_row = row[0]
            ticker = current_row.value
            industry_cell = sheet.cell(row=current_row.row, column=1)  # the instustry sector is stored in the first column
            if  industry_cell.value == 'SETOR ECONÔMICO':
                # the general industry is defined bellow the 'SETOR ECONÔMICO' header, however that header his merged and
                # occupies 2 rows, that's why there +2. Note that, until the next header, all the firms belongs to the
                # same industry
                industry = sheet.cell(row=(industry_cell.row+2), column=1).value
            if ticker != None and (len(ticker) == 4):
                #tickers.append(ticker)
                row_addr = current_row.row
                tickers[row_addr] = {'Ticker': ticker,
                                    'Trade Name': sheet.cell(row=row_addr, column=3).value.strip(),
                                     'Industry': industry.strip()}
        b3_df = DataFrame.from_dict(tickers, orient='index')
        b3_df.to_excel("B3_list.xlsx", index=False)


if __name__ == "__main__":
    plan = Plan()
    plan.organize_plan()


